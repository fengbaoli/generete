# -*- coding:utf-8 -*-
import  os,xlrd
import xlsxwriter
from openpyxl import load_workbook
import  sys
reload(sys)
sys.setdefaultencoding('utf-8')

filename ="chart_column.xlsx"

#如果存在文件先进行删除
if os.path.exists(filename):
    os.remove(filename)

def merge_xlsx(filename):
    data = xlrd.open_workbook(filename)
    wb = load_workbook(filename)
    for project_name in data.sheet_names():
        table = data.sheet_by_name(project_name)
        #sheet 总行数
        nrows = table.nrows
        #生成合并单元格列表
        cell = []
        for i in range(1,nrows):
            cell.append(table.cell(i,0).value)
        #生成没有重复的单元格列表
        cell1 = list(set(cell))
        #获取sheet单元格值的重复起始及结束值
        for key in cell1:
            counts =0
            for i in range(0,len(cell)):
                if cell[i] == key:
                    counts = counts+1
            start = cell.index(key)+2
            end = counts+start - 1
            #合并单元格
            ws = wb.active
            ws = wb[project_name]
            merge_start_row = "A"+str(start)
            merge_end_row = "A"+str(end)
            merge_cells_select = merge_start_row+":"+merge_end_row
            ws.merge_cells(merge_cells_select)
    wb.save(filename)

def generate_xlsx(xlsx_filename):
    #遍历当前目录
    items = os.listdir(".")
    #定义env文件列表
    env_filename_list = []
    #定义xlsx文件名
    workbook = xlsxwriter.Workbook(xlsx_filename)
    #筛选当前目录的env文件
    for names in items:
      if names.endswith(".env"):
        env_filename_list.append(names)
    #xlxs起始sheet
    start_sheet =1
    #从env列表获取env文件名
    for env_filename in env_filename_list:
        env_filename = env_filename
        project_name =  env_filename.split(".")[0]
        path = os.getcwd()+"/"+project_name
        data = []
        job_name = []
        key_name = []
        key_values = []
        for root, dirs, files in os.walk(path):
            for filename in files:
                #打开env文件
                env_file = open(env_filename)
                while True:
                    #一行一行读取env文件内容
                    line = env_file.readline()
                    #获取env文件key值
                    line1 = line.split("=")[0]
                    if line:
                        #打开工程文件
                        pro_file = open(path+"/"+filename)
                        while True:
                            line2 = pro_file.readline()
                            pro_key = line2.replace("$","").strip("\n")
                            if line2:
                                #两个文件内容匹配
                                if  line1 == pro_key:
                                    #获取env文件的value值
                                    values = line.split("=")[1].strip("\n")
                                    #从工程文件获取key值
                                    pro_key = line2.replace("$","").strip("\n")
                                    #写入csv文件
                                    job_name.append(filename)
                                    key_name.append(pro_key)
                                    key_values.append(values)
                            else:
                                break
                        pro_file.close()
                    else:
                        break
                env_file.close()
        #生成数据列表
        data.append(job_name)
        data.append(key_name)
        data.append(key_values)
        #print data
        #生成excel文件
        sheet="worksheet"+str(start_sheet)
        sheet = workbook.add_worksheet(project_name)
        bold = workbook.add_format({'bold': 5})
        bold.set_font_size(15)
        bold.set_bg_color('red')
        bold.set_align('center')
        bold.set_border(5)
        #第一列格式化
        column1_format = workbook.add_format()
        #字体大小
        column1_format.set_font_size(8)
        '''
        #字体颜色
        #column1_format.set_font_color('red')
        #字体斜体
        #column1_format.set_italic()
        #居中
        #column1_format.set_align('center')
        '''
        #背景颜色
        column1_format.set_bg_color('green')
        column1_format.set_border(1)
        column2_format = workbook.add_format()
        column2_format.set_font_size(8)
        column2_format.set_bg_color('yellow')
        column2_format.set_border(1)
        column3_format = workbook.add_format()
        column3_format.set_font_size(8)
        column3_format.set_bg_color('yellow')
        column3_format.set_border(1)
        headings = ['job名', '变量', '值']
        sheet.write_row('A1', headings, bold)
        sheet.set_column('A:A', 50)
        sheet.set_column('B:B', 50)
        sheet.set_column('C:C', 50)
        sheet.write_column('A2', data[0],column1_format)
        sheet.write_column('B2', data[1],column2_format)
        sheet.write_column('C2', data[2],column3_format)
        start_sheet=start_sheet+1
    workbook.close()

if __name__=="__main__":
    generate_xlsx(filename)
    #进行单元格合并
    merge_xlsx(filename)
