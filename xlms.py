# -*- coding:utf-8 -*-
import  os,xlrd
import xlsxwriter
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
import re,ConfigParser
import datetime,time
import  sys
from collections import Counter
reload(sys)
sys.setdefaultencoding('utf-8')
from debug import logdebug
log = logdebug()

#定义环境变量
filename ="chart_column.xlsx"
project_59 = ['10.249.15.59',['dpods','dpsor','dpstc']]
project_60 = ['10.249.15.60',['dpdpa']]
project_40 = ['10.249.34.40',['dpsrc','dpdsa']]
odbc_file ="/opt/IBM/InformationServer/Server/DSEngine/.odbc.ini"
host_list = ['10.249.15.59','10.249.15.60','10.249.34.40']
tmp_filename = "text.text"
merge_file = "merge_file.text"
ds_username = "dsadm"
ds_password = "dsadm"
oracle_username = "oracle"
oracle_password = "oracle"


def ssh_cmd(ip,port,username,password,cmd):
    import paramiko
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(ip,port,username,password)
    stdin, stdout, stderr = ssh.exec_command(cmd)
    return stdout.readlines()

def format_odbc_file(filename):
    host_ip = filename.split("/")[1].split("-")[0]
    cf = ConfigParser.ConfigParser()
    cf.read(filename)
    mysql_driver="libmyodbc5w.so"
    gp_driver= "VMgplm00.so"
    mysql_sections_list = []
    gp_sections_list = []
    for sections in cf.sections():
        for key_name in cf.items(sections):
            #mysql配置
            if re.findall(mysql_driver,key_name[1]):
                mysql_sections_list.append(sections)
            #GP配置
            if re.findall(gp_driver,key_name[1]):
                if re.findall("Greenplum Wire Protocol",sections):
                    pass
                else:
                    gp_sections_list.append(sections)
    mysql_list=[]
    gp_list = []
    #获取mysql的端口，用户，密码等配置
    for mysql_name  in mysql_sections_list:
        odbc_temp_list= []
        server_name = cf.get(mysql_name,"SERVER")
        port =cf.get(mysql_name,"PORT")
        database_name =cf.get(mysql_name,"DATABASE")
        odbc_temp_list.append(host_ip)
        odbc_temp_list.append(mysql_name)
        odbc_temp_list.append(server_name)
        odbc_temp_list.append(port)
        odbc_temp_list.append(database_name)
        odbc_temp_list.append("mysql")
        mysql_list.append(odbc_temp_list)
    #获取GP的端口，用户，密码等配置
    for gp_name  in gp_sections_list:
        odbc_temp_list= []
        server_name = cf.get(gp_name,"HostName")
        port =cf.get(gp_name,"PortNumber")
        database_name =cf.get(gp_name,"Database")
        odbc_temp_list.append(host_ip)
        odbc_temp_list.append(gp_name)
        odbc_temp_list.append(server_name)
        odbc_temp_list.append(port)
        odbc_temp_list.append(database_name)
        odbc_temp_list.append("gp")
        gp_list.append(odbc_temp_list)
    return  mysql_list,gp_list


def format_tns_file(filename):
    host_ip = filename.split("/")[1].split("-")[0]
    f = open(filename, "rb")
    #第一次格式化的初始列表（提取host，port，service，tnsname的列表）
    tns_list = []
    while True:
        line = f.readline()
        #定义host匹配
        host = "HOST"
        #定义service_name匹配
        service_name = "SERVICE_NAME"
        #定义port匹配
        port = "PORT"
        if line:
            #获取tnsname
            if re.match(r"^\w+",line):
                tnsname = line.replace('=','').strip()
                tns_list.append(tnsname)
            #获取host
            if re.findall(host,line):
                host = line.split("HOST")[1].split(")")[0].replace('=','').strip()
                tns_list.append(host)
            #获取port
            if re.findall(port,line):
                port= line.split(port)[1].replace(')','').replace('=','').strip()
                tns_list.append(port)
            #获取service_name
            if re.findall(service_name,line):
                s_name = line.replace('(','').replace(')','').split("=")[1]
                tns_list.append(s_name)
        else:
            break
    f.close()
    #格式化列表
    #最终格式化的列表
    tnsname_list = []
    i = 0
    while i <len(tns_list)/4:
        #一元临时列表
        t_list = []
        t_name = tns_list[i]
        t_list.append(host_ip)
        t_list.append(t_name)
        t_host = tns_list[i+1]
        t_list.append(t_host)
        t_port = tns_list[i+2]
        t_list.append(t_port)
        t_service_name = tns_list[i+3].strip()
        t_list.append(t_service_name)
        t_list.append("oracle")
        #将一元列表加入到最终列表组成二元列表
        tnsname_list.append(t_list)
        i = i + 4
    return  tnsname_list

def get_remote_tns_file(host_list):
    for host_ip in host_list:
        #oracle
        cmd ='cat .bash_profile |grep  ORACLE_HOME|grep -v PATH|awk -F"=" \'{print $2}\' '
        for oraclehome in ssh_cmd(host_ip,22,username=oracle_username,password=oracle_password,cmd=cmd):
            oracle_home = oraclehome.strip()
            if re.findall("ORACLE_BASE",oracle_home):
                base_cmd ='cat .bash_profile |grep  ORACLE_BASE|grep -v PATH|grep -v ORACLE_HOME|awk -F"=" \'{print $2}\' '
                for oraclebase in ssh_cmd(host_ip,22,'oracle','oracle',base_cmd):
                    oracle_base = oraclebase.strip()
                oracle_home=oracle_home.replace("$ORACLE_BASE",oracle_base)
        cmd_get_tns_contents = 'cat %s/network/admin/tnsnames.ora' %(oracle_home)
        #清除原来文件
        tns_filename = host_ip+"-tnsnames.ora"
        if not os.path.exists("tmp"):
            os.mkdir("tmp")
        if os.path.exists("tmp/"+tns_filename):
            os.remove("tmp/"+tns_filename)
        for tns_contents in ssh_cmd(host_ip,22,username=oracle_username,password=oracle_password,cmd=cmd_get_tns_contents):
            tns_file = open("tmp/"+tns_filename,"a+")
            tns_file.write(tns_contents)
            tns_file.close()


def get_remote_odbc_file(host_list):
    for host_ip in host_list:
        cmd_get_odbc_contents = 'cat /opt/IBM/InformationServer/Server/DSEngine/.odbc.ini'
        #清除原来文件
        odbc_filename = host_ip+"-odbc.ini"
        if not os.path.exists("tmp"):
            os.mkdir("tmp")
        if os.path.exists("tmp/"+odbc_filename):
            os.remove("tmp/"+odbc_filename)
        for odbc_contents in ssh_cmd(host_ip,22,username=ds_username,password=ds_password,cmd=cmd_get_odbc_contents):
            odbc_file = open("tmp/"+odbc_filename,"a+")
            odbc_file.write(odbc_contents)
            odbc_file.close()

def generate_file(filename,host_list,tmp_filename):
    #从远程生产tns文件
    get_remote_tns_file(host_list)
    #从远程生产odbc文件
    get_remote_odbc_file(host_list)
    #xlxs文件如果存在进行删除
    if os.path.exists(filename):
        os.remove(filename)
    #生成文件
    workbook = xlsxwriter.Workbook(filename)
    bold = workbook.add_format({'bold': 5})
    bold.set_font_size(15)
    bold.set_bg_color('red')
    bold.set_align('center')
    bold.set_border(5)
    #第一列格式化
    column1_format = workbook.add_format()
    #字体大小
    column1_format.set_font_size(8)
    #背景颜色
    column1_format.set_bg_color('green')
    column1_format.set_border(1)
    column2_format = workbook.add_format()
    column2_format.set_font_size(8)
    column2_format.set_bg_color('yellow')
    column2_format.set_border(1)
    #格式化远程文件
    tns_data = []
    mysql_data = []
    gp_data = []
    for parent,dirnames,filenames in os.walk("tmp"):
        for filename in filenames:
            endsig = os.path.splitext(filename)[1]
            if endsig == ".ora":
                tns_data.append(format_tns_file(os.path.join(parent,filename)))
            if endsig == ".ini":
                mysql_list,gp_list=format_odbc_file(os.path.join(parent,filename))
                mysql_data.append(mysql_list)
                gp_data.append(gp_list)

    if os.path.exists(tmp_filename):
        os.remove(tmp_filename)
    for project_name in ["oracle","mysql","gp"]:
        host_ip = []
        connect_name = []
        port =[]
        service_ip = []
        service_name = []
        sheet = workbook.add_worksheet(project_name)
        file = open(tmp_filename,"a+")
        if project_name == "oracle":
            for single_host_tns in tns_data:
                for single_tns in single_host_tns:
                    host_ip.append(single_tns[0])
                    connect_name.append(single_tns[1])
                    service_ip.append(single_tns[2])
                    port.append(single_tns[3])
                    service_name.append(single_tns[4])
                    file.write(single_tns[1]+":"+single_tns[2]+":"+single_tns[3]+":"+single_tns[4]+":"+single_tns[0]+":"+single_tns[5]+"\n")
        if project_name == "mysql":
            for single_host_jdbc   in mysql_data:
                for single_jdbc in single_host_jdbc:
                    host_ip.append(single_jdbc[0])
                    connect_name.append(single_jdbc[1])
                    service_ip.append(single_jdbc[2])
                    port.append(single_jdbc[3])
                    service_name.append(single_jdbc[4])
                    file.write(single_jdbc[1]+":"+single_jdbc[2]+":"+single_jdbc[3]+":"+single_jdbc[4]+":"+single_jdbc[0]+":"+single_jdbc[5]+"\n")
        if project_name == "gp":
            for single_host_gp   in gp_data:
                for single_gp in single_host_gp:
                    host_ip.append(single_gp[0])
                    connect_name.append(single_gp[1])
                    service_ip.append(single_gp[2])
                    port.append(single_gp[3])
                    service_name.append(single_gp[4])
                    file.write(single_gp[1]+":"+single_gp[2]+":"+single_gp[3]+":"+single_gp[4]+":"+single_gp[0]+":"+single_gp[5]+"\n")
        headings = ['etl主机ip','连接串', 'ip地址', '端口','数据库']
        sheet.write_row('A1', headings, bold)
        sheet.set_column('A:A', 15)
        sheet.set_column('B:B', 10)
        sheet.set_column('C:C', 20)
        sheet.set_column('D:D', 10)
        sheet.set_column('E:E', 10)
        sheet.write_column('A2', host_ip,column1_format)
        sheet.write_column('B2', connect_name,column2_format)
        sheet.write_column('C2', service_ip,column2_format)
        sheet.write_column('D2', port,column2_format)
        sheet.write_column('E2', service_name,column2_format)
    file.close()
    workbook.close()


def merge_xlsx(filename):
    data = xlrd.open_workbook(filename)
    wb = load_workbook(filename)
    for project_name in data.sheet_names():
        table = data.sheet_by_name(project_name)
        #sheet 总行数
        nrows = table.nrows
        #生成合并单元格列表
        cell = []
        for i in range(1,nrows):
            cell.append(table.cell(i,0).value)
        #生成没有重复的单元格列表
        cell1 = list(set(cell))
        #获取sheet单元格值的重复起始及结束值
        for key in cell1:
            counts =0
            for i in range(0,len(cell)):
                if cell[i] == key:
                    counts = counts+1
            start = cell.index(key)+2
            end = counts+start - 1
            #合并单元格
            ws = wb.active
            ws = wb[project_name]
            merge_start_row = "A"+str(start)
            merge_end_row = "A"+str(end)
            merge_cells_select = merge_start_row+":"+merge_end_row
            ws.merge_cells(merge_cells_select)
    wb.save(filename)

def generate_xlsx(xlsx_filename):
    #遍历当前目录
    items = os.listdir(".")
    #定义env文件列表
    env_filename_list = []
    #定义xlsx文件名
    workbook = xlsxwriter.Workbook(xlsx_filename)
    #筛选当前目录的env文件
    for names in items:
      if names.endswith(".env"):
        env_filename_list.append(names)
    #xlxs起始sheet
    start_sheet =1
    #从env列表获取env文件名
    for env_filename in env_filename_list:
        env_filename = env_filename
        project_name =  env_filename.split(".")[0]
        path = os.getcwd()+"/"+project_name
        data = []
        job_name = []
        key_name = []
        key_values = []
        for root, dirs, files in os.walk(path):
            for filename in files:
                #打开env文件
                env_file = open(env_filename)
                while True:
                    #一行一行读取env文件内容
                    line = env_file.readline()
                    #获取env文件key值
                    line1 = line.split("=")[0]
                    if line:
                        #打开工程文件
                        pro_file = open(path+"/"+filename)
                        while True:
                            line2 = pro_file.readline()
                            pro_key = line2.replace("$","").strip("\n")
                            if line2:
                                #两个文件内容匹配
                                if  line1 == pro_key:
                                    #获取env文件的value值
                                    values = line.split("=")[1].strip("\n")
                                    #从工程文件获取key值
                                    pro_key = line2.replace("$","").strip("\n")
                                    #写入csv文件
                                    job_name.append(filename)
                                    key_name.append(pro_key)
                                    key_values.append(values)
                            else:
                                break
                        pro_file.close()
                    else:
                        break
                env_file.close()
        #生成数据列表
        data.append(job_name)
        data.append(key_name)
        data.append(key_values)
        #print data
        #生成excel文件
        sheet="worksheet"+str(start_sheet)
        sheet = workbook.add_worksheet(project_name)
        bold = workbook.add_format({'bold': 5})
        bold.set_font_size(15)
        bold.set_bg_color('red')
        bold.set_align('center')
        bold.set_border(5)
        #第一列格式化
        column1_format = workbook.add_format()
        #字体大小
        column1_format.set_font_size(8)
        #背景颜色
        column1_format.set_bg_color('green')
        column1_format.set_border(1)
        column2_format = workbook.add_format()
        column2_format.set_font_size(8)
        column2_format.set_bg_color('yellow')
        column2_format.set_border(1)
        column4_format = workbook.add_format()
        column4_format.set_font_size(8)
        column4_format.set_bg_color('gray')
        column4_format.set_border(1)
        headings = ['job名', '变量', '值','ip','端口','service_name','etl主机ip地址']
        sheet.write_row('A1', headings, bold)
        sheet.set_column('A:A', 32)
        sheet.set_column('B:B', 25)
        sheet.set_column('C:C', 35)
        sheet.set_column('D:D', 15,column4_format)
        sheet.set_column('E:E', 5,column4_format)
        sheet.set_column('F:F', 18,column4_format)
        sheet.set_column('G:G', 40,column4_format)
        sheet.write_column('A2', data[0],column1_format)
        sheet.write_column('B2', data[1],column2_format)
        sheet.write_column('C2', data[2],column2_format)
        start_sheet=start_sheet+1
    workbook.close()

def modify_xlxs(filename,text_filename):
    data = xlrd.open_workbook(filename)
    wb = load_workbook(filename)
    #将text_filename中的内容加载到data的列表中
    data_list = []
    file = open(text_filename)
    while True:
        line = file.readline()
        if line:
            data_list.append(line.strip())
        else:
            break
    file.close()

    for project_name in data.sheet_names():
        table = data.sheet_by_name(project_name)
        ws = wb.get_sheet_by_name(project_name)
        #sheet 总行数
        nrows = table.nrows
        for i in range(1,nrows):
            cell = table.cell(i,2).value
            #遍历data_list
            for merge_values in data_list:
                merge_sp=merge_values.split(":")
                #处理gp和mysql
                if merge_sp[5].strip() == "gp" or merge_sp[5].strip() == "mysql":
                    if merge_sp[0].strip() == cell:
                        connect_name = merge_sp[1]
                        port = merge_sp[2]
                        database = merge_sp[3]
                        etl_ip = merge_sp[4]
                        #写入xls文件conectname =(i,3),port =(i,4),database=(i,5),etl_ip=(i,6)
                        ws.cell(row=int(i+1),column=4).value = connect_name
                        ws.cell(row=int(i+1),column=5).value = port
                        ws.cell(row=int(i+1),column=6).value = database
                        ws.cell(row=int(i+1),column=7).value = etl_ip
                #处理oracle
                if merge_sp[5].strip() == "oracle":
                    if merge_sp[0].strip() == cell:
                        connect_name = merge_sp[1]
                        port = merge_sp[2]
                        database = merge_sp[3]
                        etl_ip = merge_sp[4]
                        if project_name in ['dpods','dpsor','dpstc'] and etl_ip.strip() == "10.249.15.59":
                            ws.cell(row=int(i+1),column=4).value = connect_name
                            ws.cell(row=int(i+1),column=5).value = port
                            ws.cell(row=int(i+1),column=6).value = database
                            ws.cell(row=int(i+1),column=7).value = etl_ip
                        if  project_name in ['dpsrc','dpdsa'] and etl_ip.strip() == "10.249.34.40":
                            ws.cell(row=int(i+1),column=4).value = connect_name
                            ws.cell(row=int(i+1),column=5).value = port
                            ws.cell(row=int(i+1),column=6).value = database
                            ws.cell(row=int(i+1),column=7).value = etl_ip           
    wb.save(filename)

def merge_mysql_gp_text_file(tmp_filename,merge_file):
    if not os.path.exists(tmp_filename):
        print("Error,%s 文件不存在") % (tmp_filename)
    if os.path.exists(merge_file):
        os.remove(merge_file)
    file = open(tmp_filename)
    merge_file = open(merge_file,"a+")
    merge_list1 = []
    merge_list2 = []
    while True:
        line = file.readline()
        if line:
            #flag获取读取的一行内容的最后一个切片的值即为oracle或者mysql或gp
            flag=line.split(":")[5].strip()
            #对读取的一行数据按照冒号进行切片
            line_sp = line.split(":")
            #如果最后一个切片的内容为gp，然后就加入到列表中，merge_list1为就有原纪录的的列表，merge_list2为去重记录的列表
            if flag == "gp" :
                merge_str1 = line_sp[0]+":"+line_sp[1]+":"+line_sp[2]+":"+line_sp[3]+":gp"
                merge_list1.append(merge_str1)
                merge_list1 = list(set(merge_list1))
                merge_list2.append(line)
            if flag ==  "mysql":
                merge_str1 = line_sp[0]+":"+line_sp[1]+":"+line_sp[2]+":"+line_sp[3]+":mysql"
                merge_list1.append(merge_str1)
                merge_list1 = list(set(merge_list1))
                merge_list2.append(line)
            if flag == "oracle":
                merge_file.write(line)
        else:
            break
    #从列表找出相同记录，如果存在重复记录，就合切片的下标为4的主机地址记录，以"/"分隔,然后写入文件
    for contents1 in merge_list1:
        data = []
        contents1_sp = contents1.split(":")
        for contents2 in merge_list2:
            contents2_sp = contents2.split(":")
            str1 = contents2_sp[0]+":"+contents2_sp[1]+":"+contents2_sp[2]+":"+contents2_sp[3]
            str2 = contents1_sp[0]+":"+contents1_sp[1]+":"+contents1_sp[2]+":"+contents1_sp[3]
            if str2 == str1 :
                data.append(contents2.split(":")[4])
        #list转str
        str_convert = '/'.join(data)
        #拼接为str，格式为连接串：ip：port：database：etl主机地址（重复的以"/"分隔进行合并）：标志
        str3 = contents1_sp[0]+":"+contents1_sp[1]+":"+contents1_sp[2]+":"+contents1_sp[3]+":"+str_convert+":"+contents1_sp[4]+"\n"
        merge_file.write(str3)
    merge_file.close()
    file.close()
    #删除临时文件tmp_filename
    if os.path.exists(tmp_filename):
        os.remove(tmp_filename)


def main():
    #从tns,odbc获取相关的值，并生成jdbc.xlsx及text.text临时文件
    message = "脚本开始运行......"
    log.loginfo(message)
    print("脚本开始运行......")
    message="从tns,odbc获取相关的值，并生成jdbc.xlsx及text.text临时文件开始"
    print("%s: 从tns,odbc获取相关的值，并生成jdbc.xlsx及text.text临时文件开始") % (time.strftime("%Y-%m-%d %H:%M:%S"))
    log.loginfo(message)
    generate_file(host_list=host_list,filename="jdbc.xlsx",tmp_filename=tmp_filename)
    message = "开始合并text.text临时文件的host地址"
    print("%s: 开始合并text.text临时文件的host地址") % (time.strftime("%Y-%m-%d %H:%M:%S"))
    log.loginfo(message)
    merge_mysql_gp_text_file(tmp_filename,merge_file)
    message="合并text.text临时文件的host地址结束"
    log.loginfo(message)
    message="从tns,odbc获取相关的值，并生成jdbc.xlsx及text.text临时文件结束"
    log.loginfo(message)
    print("%s: 合并text.text临时文件的host地址结束") % (time.strftime("%Y-%m-%d %H:%M:%S"))
    print("%s: 从tns,odbc获取相关的值，并生成jdbc.xlsx及text.text临时文件结束") % (time.strftime("%Y-%m-%d %H:%M:%S"))
    #从当前目录生成临时的chart_column.xlsx文件并对相同单元格合并
    message = "从当前目录生成临时的chart_column.xlsx文件并对相同单元格合并开始"
    log.loginfo(message)
    print("%s: 从当前目录生成临时的chart_column.xlsx文件并对相同单元格合并开始") % (time.strftime("%Y-%m-%d %H:%M:%S"))
    if os.path.exists(filename):
        os.remove(filename)
    generate_xlsx(filename)
    message = "生成chart_column.xlsx文件结束,开始进行单元格合并"
    log.loginfo(message)
    print("%s: 生成chart_column.xlsx文件结束,开始进行单元格合并") % (time.strftime("%Y-%m-%d %H:%M:%S"))
    #进行单元格合并
    merge_xlsx(filename)
    message = "文件chart_column.xlsx单元格合并结束"
    log.loginfo(message)
    print("%s: 文件chart_column.xlsx单元格合并结束") % (time.strftime("%Y-%m-%d %H:%M:%S"))
    message = "文件chart_column.xlsx匹配tns,odbc开始"
    log.loginfo(message)
    print("%s: 文件chart_column.xlsx匹配tns,odbc开始") % (time.strftime("%Y-%m-%d %H:%M:%S"))
    modify_xlxs(filename,merge_file)
    message = "文件chart_column.xlsx匹配tns,odbc结束"
    log.loginfo(message)
    print("%s: 文件chart_column.xlsx匹配tns,odbc结束") % (time.strftime("%Y-%m-%d %H:%M:%S"))
    message = "任务成功结束......"
    log.loginfo(message)
    print ("任务成功结束......")

if __name__=="__main__":
    main()
